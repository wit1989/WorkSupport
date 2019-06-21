#!/usr/bin/env python3
# -*- coding:utf-8 -*-


import tkinter as tk
from tkinter import ttk
import tkinter.filedialog as filedialog
import tkinter.messagebox
import os, xlrd, xlwt, openpyxl
from openpyxl.utils import column_index_from_string

myfont = ('Monaco', 12)

"""切换函数"""


def data_comparison():
    global switch
    switch.pack_forget()
    data_comparison_frm0.pack(fill='both', expand=1, side=tk.BOTTOM)
    data_comparison_input.delete(0, 'end')
    data_comparison_input.insert(0, '此处输入要参与对比的列，无间隔（双击清空）')
    switch = data_comparison_frm0


def merge_excel():
    global switch
    switch.pack_forget()
    merge_excel_frm0.pack(fill='both', expand=1)
    merge_excel_input.delete(0, 'end')
    merge_excel_input.insert(0, '此处输入表头行数（双击清空）')
    switch = merge_excel_frm0


def non_standard():
    global switch
    switch.pack_forget()
    non_standard_frm0.pack(fill='both', expand=1)
    switch = non_standard_frm0


def timing_shutdown():
    pass


def shutdown_iu():
    switch.pack_forget()


"""功能函数"""

"""数据对比"""
def to_str(string):
    if isinstance(string, float):
        return str(int(string))
    else:
        return string


def filename():
    fname.set('')
    address = filedialog.askopenfilename()
    if address:
        fname.set(address)

    if os.path.basename(address) and not (
            os.path.basename(address).endswith('.xls') or os.path.basename(address).endswith('.xlsx')):
        tkinter.messagebox.showinfo('提示', '所选文件格式不正确！')


def comparison_xls(address):
    """xls格式，取两个sheet后进行对比，返回对比结果error_data"""
    columns = data_comparison_input.get()
    wb = xlrd.open_workbook(address)
    sheetone = wb.sheet_by_index(0)
    sheettwo = wb.sheet_by_index(1)

    sheetone_data = []
    sheettwo_data = []
    error_data = []
    for i in range(0, sheetone.nrows):
        sheetone_column = []
        for co in columns:
            sheetone_column.append(to_str(sheetone.cell(i, column_index_from_string(co)).value))
        sheetone_data.append(sheetone_column)
    for ii in range(0, sheettwo.nrows):
        sheettwo_column = []
        for ct in columns:
            sheettwo_column.append(to_str(sheettwo.cell(ii, column_index_from_string(ct)).value))
        sheettwo_data.append(sheettwo_column)

    num = 1
    for soc in range(len(sheetone_data)):
        error_column = []
        if sheetone_data[soc] not in sheettwo_data:
            error_column.append(num)
            error_column.append('在表1不在表2')
            error_column.append(soc + 1)
            error_column.append(sheetone_data[soc])
            error_data.append(error_column)
            num += 1
    for stc in range(len(sheettwo_data)):
        error_column = []
        if sheettwo_data[stc] not in sheetone_data:
            error_column.append(num)
            error_column.append('在表2不在表1')
            error_column.append(stc + 1)
            error_column.append(sheettwo_data[stc])
            error_data.append(error_column)
            num += 1
    print(error_data)
    return error_data


def comparison_xlsx(address):
    """xlsx格式，取两个sheet后进行对比，返回对比结果error_data"""
    columns = data_comparison_input.get()
    wb = openpyxl.load_workbook(address)
    sheetone = wb.worksheets[0]
    sheettwo = wb.worksheets[1]

    # 生成sheet1和sheet2 所有数据的列表 [[第一行数据拼接],[第二行数据拼接],...]
    sheetone_data = []
    sheettwo_data = []
    error_data = []
    # 注：openpyxl的row和column是从1开始，xlrd是从0开始
    for i in range(1, sheetone.max_row + 1):
        sheetone_column = []
        for co in columns:
            sheetone_column.append(sheetone.cell(row=i, column=column_index_from_string(co)).value)
        sheetone_data.append(sheetone_column)

    for ii in range(1, sheettwo.max_row + 1):
        sheettwo_column = []
        for ct in columns:
            sheettwo_column.append(sheettwo.cell(row=ii, column=column_index_from_string(ct)).value)
        sheettwo_data.append(sheettwo_column)

    num = 1
    for soc in range(len(sheetone_data)):
        error_column = []
        if sheetone_data[soc] not in sheettwo_data:
            error_column.append(num)
            error_column.append('在表1不在表2')
            error_column.append(soc + 1)
            error_column.append(sheetone_data[soc])
            error_data.append(error_column)
            num += 1
    for stc in range(len(sheettwo_data)):
        error_column = []
        if sheettwo_data[stc] not in sheetone_data:
            error_column.append(num)
            error_column.append('在表2不在表1')
            error_column.append(stc + 1)
            error_column.append(sheettwo_data[stc])
            error_data.append(error_column)
            num += 1
    print(error_data)

    return error_data

    # 输出结果到新的sheet中
    # newsheet = '对比结果'
    # wb.create_sheet(newsheet)
    # resultsheet = wb.worksheets[-1]
    # resultsheet['A1'] = '错误'
    # resultsheet['B1'] = '行数'
    # resultsheet['C1'] = '字段内容'
    #
    # num = 2
    # for soc in range(len(sheetone_data)):
    #     if sheetone_data[soc] not in sheettwo_data:
    #         resultsheet['A' + str(num)] = '在sheet1中不在sheet2中'
    #         resultsheet['B' + str(num)] = soc + 1
    #         resultsheet['C' + str(num)] = str(sheetone_data[soc])
    #         print('表' + wb.sheetnames[0] + '  第' + str(soc + 1) + '行  ' + str(sheetone_data[soc]) + '  不在表' +
    #               wb.sheetnames[1] + '中')
    #         num += 1
    #
    # for stc in range(len(sheettwo_data)):
    #     if sheettwo_data[stc] not in sheetone_data:
    #         resultsheet['A' + str(num)] = '在sheet2中不在sheet1中'
    #         resultsheet['B' + str(num)] = stc + 1
    #         resultsheet['C' + str(num)] = str(sheettwo_data[stc])
    #         print('表' + wb.sheetnames[1] + '  第' + str(stc + 1) + '行  ' + str(sheettwo_data[stc]) + '  不在表' +
    #               wb.sheetnames[0] + '中')
    #         num += 1
    #
    # wb.save(filename=name[:-5] + '(对比结果)' + '.xlsx')


def begin_comparison():
    address = fname.get()
    fn = os.path.basename(fname.get())

    """清空Treeview中的内容"""
    x = data_comparison_tv.get_children()
    for item in x:
        data_comparison_tv.delete(item)

    """将error_data放入Treeview中"""
    if fn.endswith('.xls'):
        # print(comparison_xls(address))
        if comparison_xls(address):
            for i in comparison_xls(address):
                data_comparison_tv.insert('', i[0], text=i[0], values=i[0:])
        else:
            tkinter.messagebox.showinfo("提示", "数据完全一致！")
    elif fn.endswith('.xlsx'):
        if comparison_xlsx(address):
            for i in comparison_xlsx(address):
                data_comparison_tv.insert('', i[0], text=i[0], values=i[0:])
        else:
            tkinter.messagebox.showinfo("提示", "数据完全一致！")


"""合并Excel"""
def filedir():
    global path
    fdir.set('')  # 清空文本框里内容
    flist.set('')
    path = filedialog.askdirectory()
    if path:
        fdir.set(path)
    flist.set(os.listdir(path))


def agg_excel(me_input, fpath):
    # global merge_excel_input, path
    if isinstance(int(me_input.get()), int):
        sheethead = int(me_input.get())
    else:
        sheethead = 0
    # path = '/Users/cbowen/downloads/分院上报材料/all'
    os.chdir(fpath)

    all_wb = xlwt.Workbook(encoding="utf-8", style_compression=0)
    all_sheet = all_wb.add_sheet("数据汇总")

    max_row = 0
    for file in data_files():
        print(file)
        sub_wb = xlrd.open_workbook(file)
        sub_sheet = sub_wb.sheet_by_index(0)

        for row in range(sub_sheet.nrows - sheethead):
            for column in range(sub_sheet.ncols):
                all_sheet.write(max_row + row, column, sub_sheet.cell_value(row + sheethead, column))
            row += 1
        max_row += sub_sheet.nrows - sheethead

    last_name = file_name('alldata.xls')
    all_wb.save(last_name)

    tkinter.messagebox.showinfo('提示', '合并完成,请查看《' + str(last_name) + '》')


def input_num(en=None):
    """清空输入框中的提示文字，各功能通用"""
    merge_excel_input.delete('0', 'end')
    data_comparison_input.delete('0', 'end')


def file_name(name):
    """保存时重名检测"""
    if (name) not in os.listdir('.'):  # 如果本地目录下有重名文件，则在文件名后面加"(n)"，n顺次加1
        even_name = name
    else:
        file_num = 1
        while True:
            dot = name.rfind('.')  # 查询文件名中最后一个'.'的索引
            head = name[:dot]
            tail = name[dot:]

            new_name = head + '(' + str(file_num) + ')' + tail
            if new_name not in os.listdir('.'):
                even_name = new_name
                break
            else:
                file_num += 1
    return even_name


def data_files():
    """提取当前目录下需要合并的文件名，放入列表中"""
    dadafiles = []
    for file in os.listdir('.'):
        if (file[-4:] == '.xls' or file[-5:] == '.xlsx') and file[:7] != 'alldata':
            dadafiles.append(file)
    return dadafiles


def treeview_sort_column(tv, col, reverse):
    """Treeview点击字段排序"""
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    l.sort(key=lambda t: (t[0]), reverse=reverse)

    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


"""创建GUI"""
root = tk.Tk()
root.title('Work Support')
width, heigth = 500, 400
root.geometry(
    '%dx%d+%d+%d' % (width, heigth, (root.winfo_screenwidth() - width) / 2, (root.winfo_screenheight() - heigth) / 2))

"""创建顶级菜单"""
menubar = tk.Menu(root)
root.config(menu=menubar)
excelmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='Excel操作', menu=excelmenu)
datamenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='数据格式', menu=datamenu)
osmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='系统操作', menu=osmenu)
wxmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='微信操作', menu=wxmenu)
othermenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='其他操作', menu=othermenu)

"""创建下拉菜单以及对应功能的底层容器"""
excelmenu.add_command(label='数据对比', command=data_comparison)
data_comparison_frm0 = tk.Frame(root)
excelmenu.add_command(label='合并Excel', command=merge_excel)
merge_excel_frm0 = tk.Frame(root)
switch = merge_excel_frm0
excelmenu.add_command(label='非标数据合并', command=data_comparison)
non_standard_frm0 = tk.Frame(root)
datamenu.add_command(label='DBF4', command=data_comparison)
datamenu.add_command(label="in('')", command=data_comparison)
osmenu.add_command(label='定时关机', command=timing_shutdown)
osmenu.add_command(label='定时关闭IU', command=shutdown_iu)
wxmenu.add_command(label='天气预报', command=timing_shutdown)
wxmenu.add_command(label='获取群成员', command=shutdown_iu)
wxmenu.add_command(label='控制摄像头', command=shutdown_iu)

"""创建功能控件"""

"""数据对比"""
data_comparison_frm1 = tk.Frame(data_comparison_frm0)
data_comparison_frm1.pack()
data_comparison_text = tk.Label(data_comparison_frm1,
                                text="\n提示:\n1、将要对比的两个表放在同一个excel中的前两位；\n2、两个表的字段顺序要相同。\n",
                                font=myfont, justify='left').pack()
fname = tk.StringVar()
data_comparison_ent = tk.Entry(data_comparison_frm1,
                               width=40, font=myfont, textvariable=fname).pack(fill=tk.X, side=tk.LEFT)
data_comparison_button = tk.Button(data_comparison_frm1,
                                   width=10, text='选择文件', font=myfont, command=filename).pack(fill=tk.X, side=tk.LEFT)
data_comparison_frm2 = tk.Frame(data_comparison_frm0)
data_comparison_frm2.pack()
data_comparison_input = tk.Entry(data_comparison_frm2, width=40, font=myfont)
data_comparison_input.pack(fill=tk.X, side=tk.LEFT)
data_comparison_input.bind('<Double-1>', input_num)
data_comparison_button2 = tk.Button(data_comparison_frm2, width=10, text='开始对比', font=myfont,
                                    command=begin_comparison).pack(fill=tk.X, side=tk.LEFT)

listhead = ('序号', '问题', '第几行', '内容')
data_comparison_lf = tk.Frame(data_comparison_frm0, pady=20)
data_comparison_lf.pack()
data_comparison_tv = ttk.Treeview(data_comparison_lf, height=15, show="headings", columns=listhead)
for col in listhead:
    data_comparison_tv.heading(col, text=col, command=lambda c=col: treeview_sort_column(data_comparison_tv, c, False))

data_comparison_tv.pack(side=tk.LEFT)
data_comparison_tv.heading('序号', text='序号')
data_comparison_tv.heading('问题', text='问题')  # 设置字段的显示名称
data_comparison_tv.heading('第几行', text='第几行')
data_comparison_tv.heading('内容', text='内容')
data_comparison_tv.column('序号', width=40)
data_comparison_tv.column('问题', width=100)
data_comparison_tv.column('第几行', width=40)
data_comparison_tv.column('内容', width=300)
vbar = ttk.Scrollbar(data_comparison_lf, orient=tk.VERTICAL, command=data_comparison_tv.yview)  # 设置滚动条控件
vbar.pack(side=tk.LEFT, fill='y')  # 滚动条控件与被控制的Treeview同在一个容器中，并列放置，纵向填充
data_comparison_tv.configure(yscrollcommand=vbar.set)

"""合并Excel"""
merge_excel_frm1 = tk.Frame(merge_excel_frm0)
merge_excel_frm1.pack(pady=10)
merge_excel_text = tk.Label(merge_excel_frm1,
                            text="\n提示:\n1、将需要合并的文件放在同一文件夹下；\n2、数据表放在sheet1；\n3、文件表头行数要相等。\n",
                            font=myfont, justify='left').pack()
fdir = tk.StringVar()
merge_excel_ent = tk.Entry(merge_excel_frm1, width=40, font=myfont, textvariable=fdir).pack(fill=tk.X, side=tk.LEFT)
merge_excel_button = tk.Button(merge_excel_frm1,
                               width=10, text='选择文件夹', font=myfont, command=filedir).pack(fill=tk.X, side=tk.LEFT)
flist = tk.StringVar()
# flist.set((''))
merge_excel_listbox = tk.Listbox(merge_excel_frm0, listvariable=flist).pack(fill='both', expand=1)
merge_excel_frm2 = tk.Frame(merge_excel_frm0)
merge_excel_frm2.pack()
merge_excel_input = tk.Entry(merge_excel_frm2, width=40, font=myfont)
merge_excel_input.pack(fill=tk.X, side=tk.LEFT, pady=10)
merge_excel_input.bind('<Double-1>', input_num)
merge_excel_button2 = tk.Button(merge_excel_frm2, width=10, text='开始合并', font=myfont,
                                command=lambda: agg_excel(merge_excel_input, path)).pack(fill=tk.X, side=tk.LEFT)

root.mainloop()
