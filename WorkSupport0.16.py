#!/usr/bin/env python3
# -*- coding:utf-8 -*-


import tkinter as tk
from tkinter import ttk
import tkinter.filedialog as filedialog
import tkinter.messagebox
import os, xlrd, xlwt, openpyxl, itchat, re, time, cv2
from openpyxl.utils import column_index_from_string, get_column_letter

myfont = ('Monaco', 12)
myfont2 = ('Monaco', 9)
myfont3 = ('Microsoft YaHei', 20)

"""
切换函数
设定switch参数标示当前界面，选择功能后隐藏当前界面，生成对应界面
"""

def data_comparison():
    global switch
    switch.pack_forget()
    data_comparison_frm0.pack(fill='both', expand=1, side=tk.BOTTOM)
    data_comparison_input.delete(0, 'end')  # 清空Entry控件
    data_comparison_input.insert(0, '此处输入要参与对比的列，无间隔（双击清空）')  # 向Entry控件中插入提示文字
    switch = data_comparison_frm0


def merge_excel():
    global switch
    switch.pack_forget()
    merge_excel_frm0.pack(fill='both', expand=1)
    merge_excel_input.delete(0, 'end')
    merge_excel_input.insert(0, '此处输入表头行数（双击清空）')
    switch = merge_excel_frm0


def non_standard():
    global switch
    switch.pack_forget()
    # non_standard_frm0.pack(fill='both', expand=1)
    switch = non_standard_frm0


def format_DBF():
    global switch
    switch.pack_forget()
    format_DBF_frm0.pack(fill='both', expand=1)
    switch = format_DBF_frm0


def format_in():
    global switch
    switch.pack_forget()
    format_in_frm0.pack(fill='both', expand=1)
    format_in_input.delete(0, 'end')  # 清空Entry控件
    format_in_input.insert(0, '格式"A1:A10,C1:C10,F1:F10"（双击清空）')  # 向Entry控件中插入提示文字
    switch = format_in_frm0


def timing_shutdown():
    switch.pack_forget()


def shutdown_iu():
    switch.pack_forget()


def wchat_getlist():
    global switch
    switch.pack_forget()
    wchat_getlist_frm0.pack(fill='both', expand=1)
    switch = wchat_getlist_frm0


def wchat_camera():
    global switch
    switch.pack_forget()
    wchat_camera_frm0.pack(fill='both', expand=1)
    switch = wchat_camera_frm0


def logout_wchat():
    """登陆之后程序锁，不能退出"""
    global switch
    switch.pack_forget()
    itchat.logout()


def about_this():
    """关于"""
    global switch
    switch.pack_forget()
    about_this_frm0.pack(fill='both', expand=1)
    switch = about_this_frm0


def help_for():
    """帮助"""
    global switch
    switch.pack_forget


"""功能函数"""


def get_data_from_excel(dir, choose, row=1):
    """
    从excel中获取指定数据
    dir为文件名路径
    choose为指定获取数据的范围，如果未指定范围，即choose为空则获取全部数据
    row=1为默认方式，以行为单位获取数据，否则以单元格为单位获取
    """
    pass


def input_num(en=None):
    """清空输入框中的提示文字，各功能通用"""
    data_comparison_button3.configure(state='disable')
    merge_excel_input.delete('0', 'end')
    data_comparison_input.delete('0', 'end')
    format_in_input.delete('0', 'end')


def file_name(name):
    """保存时重名检测"""
    if (name) not in os.listdir('.'):  # 如果本地目录下有重名文件，则在文件名后面加"(n)"，n顺次加1
        even_name = name
    else:
        file_num = 1
        while True:
            dot = name.rfind('.')  # 查询文件名中最后一个'.'的索引
            head = name[:dot]
            tail = name[dot:]

            new_name = head + '(' + str(file_num) + ')' + tail
            if new_name not in os.listdir('.'):
                even_name = new_name
                break
            else:
                file_num += 1
    return even_name


def treeview_sort_column(tv, col, reverse):
    """Treeview点击字段排序"""
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    l.sort(key=lambda t: (t[0]), reverse=reverse)

    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


def del_emoji(delstr, restr=''):
    """过滤emoji表情"""
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(restr, delstr)


"""数据对比"""


def to_str(string):
    """将float转换成str"""
    if isinstance(string, float):
        return str(int(string))
    else:
        return string


def filename():
    """判断选择的文件是否为Excel"""
    fname.set('')
    address = filedialog.askopenfilename()
    if address:
        fname.set(address)

    if os.path.basename(address) and not (
            os.path.basename(address).endswith('.xls') or os.path.basename(address).endswith('.xlsx')):
        tkinter.messagebox.showinfo('提示', '所选文件格式不正确！')


def comparison_xls(address):
    """xls格式，取两个sheet后进行对比，返回对比结果error_data"""
    columns = data_comparison_input.get()
    wb = xlrd.open_workbook(address)
    sheetone = wb.sheet_by_index(0)
    sheettwo = wb.sheet_by_index(1)

    sheetone_data = []
    sheettwo_data = []
    error_data = []
    for i in range(0, sheetone.nrows):
        sheetone_column = []
        for co in columns:
            sheetone_column.append(to_str(sheetone.cell(i, column_index_from_string(co)).value))
        sheetone_data.append(sheetone_column)
    for ii in range(0, sheettwo.nrows):
        sheettwo_column = []
        for ct in columns:
            sheettwo_column.append(to_str(sheettwo.cell(ii, column_index_from_string(ct)).value))
        sheettwo_data.append(sheettwo_column)

    num = 1
    for soc in range(len(sheetone_data)):
        error_column = []
        if sheetone_data[soc] not in sheettwo_data:
            error_column.append(num)
            error_column.append('在表1不在表2')
            error_column.append(soc + 1)
            error_column.append(sheetone_data[soc])
            error_data.append(error_column)
            num += 1
    for stc in range(len(sheettwo_data)):
        error_column = []
        if sheettwo_data[stc] not in sheetone_data:
            error_column.append(num)
            error_column.append('在表2不在表1')
            error_column.append(stc + 1)
            error_column.append(sheettwo_data[stc])
            error_data.append(error_column)
            num += 1
    # print(error_data)
    return error_data


def comparison_xlsx(address):
    """xlsx格式，取两个sheet后进行对比，返回对比结果error_data"""
    columns = data_comparison_input.get()
    wb = openpyxl.load_workbook(address)
    sheetone = wb.worksheets[0]
    sheettwo = wb.worksheets[1]

    # 生成sheet1和sheet2 所有数据的列表 [[第一行数据拼接],[第二行数据拼接],...]
    sheetone_data = []
    sheettwo_data = []
    error_data = []
    # 注：openpyxl的row和column是从1开始，xlrd是从0开始
    for i in range(1, sheetone.max_row + 1):
        sheetone_column = []
        for co in columns:
            sheetone_column.append(sheetone.cell(row=i, column=column_index_from_string(co)).value)
        sheetone_data.append(sheetone_column)

    for ii in range(1, sheettwo.max_row + 1):
        sheettwo_column = []
        for ct in columns:
            sheettwo_column.append(sheettwo.cell(row=ii, column=column_index_from_string(ct)).value)
        sheettwo_data.append(sheettwo_column)

    num = 1
    for soc in range(len(sheetone_data)):
        error_column = []
        if sheetone_data[soc] not in sheettwo_data:
            error_column.append(num)
            error_column.append('在表1不在表2')
            error_column.append(soc + 1)
            error_column.append(sheetone_data[soc])
            error_data.append(error_column)
            num += 1
    for stc in range(len(sheettwo_data)):
        error_column = []
        if sheettwo_data[stc] not in sheetone_data:
            error_column.append(num)
            error_column.append('在表2不在表1')
            error_column.append(stc + 1)
            error_column.append(sheettwo_data[stc])
            error_data.append(error_column)
            num += 1
    # print(error_data)
    return error_data


def begin_comparison(ev=None):
    """开始对比按键"""
    address = fname.get()
    fn = os.path.basename(fname.get())

    """清空Treeview中的内容"""
    x = data_comparison_tv.get_children()
    for item in x:
        data_comparison_tv.delete(item)

    """将error_data放入Treeview中"""
    if fn.endswith('.xls'):
        # print(comparison_xls(address))
        if comparison_xls(address):
            for i in comparison_xls(address):
                data_comparison_tv.insert('', i[0], text=i[0], values=i[0:])
        else:
            tkinter.messagebox.showinfo("提示", "数据完全一致！")
    elif fn.endswith('.xlsx'):
        if comparison_xlsx(address):
            for i in comparison_xlsx(address):
                data_comparison_tv.insert('', i[0], text=i[0], values=i[0:])
        else:
            tkinter.messagebox.showinfo("提示", "数据完全一致！")
    data_comparison_button3.configure(state='normal')


def button3_disable(ev=None):
    data_comparison_button3.configure(state='disable')


def write_in_excel(data, original_name):
    wb = openpyxl.Workbook()
    resultsheet = wb.active
    resultsheet.title = "对比结果"

    field = ['序号', '问题', '第几行']
    for col in data_comparison_input.get():
        field.append(col + "列数据")
    for c in range(len(field)):
        resultsheet[str(get_column_letter(c + 1)) + "1"] = field[c]

    for row in range(len(data)):
        resultsheet["A" + str(row + 2)] = data[row][0]
        resultsheet["B" + str(row + 2)] = data[row][1]
        resultsheet["C" + str(row + 2)] = data[row][2]
        for cc in range(len(field) - 3):
            resultsheet[str(get_column_letter(cc + 4)) + str(row + 2)] = data[row][-1][cc]

    last_name = file_name(original_name + '(对比结果)' + '.xlsx')
    wb.save(filename=last_name)
    tkinter.messagebox.showinfo('提示', '输出完成,请查看《' + str(last_name) + '》')


def output_sheet():
    address = fname.get()
    fn = os.path.basename(fname.get())
    if fn.endswith('.xls'):
        if comparison_xls(address):
            write_in_excel(comparison_xls(address), fn[:-4])
        else:
            tkinter.messagebox.showinfo("提示", "数据完全一致！")
    elif fn.endswith('.xlsx'):
        if comparison_xlsx(address):
            write_in_excel(comparison_xlsx(address), fn[:-5])
        else:
            tkinter.messagebox.showinfo("提示", "无数据")


"""合并Excel"""


def filedir():
    """将目录地址和目录中的文件写入Entry和Listbox控件"""
    global path
    fdir.set('')  # 清空文本框里内容
    flist.set('')
    path = filedialog.askdirectory()  # 选择目录
    if path:
        fdir.set(path)
    flist.set(os.listdir(path))


def agg_excel(me_input, fpath):
    """开始合并，两个参数：表头行数和目录地址"""
    # global merge_excel_input, path
    if isinstance(int(me_input.get()), int):
        sheethead = int(me_input.get())
    else:
        sheethead = 0
    # path = '/Users/cbowen/downloads/分院上报材料/all'
    os.chdir(fpath)

    all_wb = openpyxl.Workbook()
    all_sheet = all_wb.active
    all_sheet.title = "数据汇总"

    for file in data_files():
        # 注：新建的空excel，用max_row获取最大行数 也会返回1。
        max_row = all_sheet.max_row
        choose_xls_xlsx(max_row, all_sheet, file, sheethead)

    all_sheet.cell(1, 1, "全部数据：")
    last_name = file_name('alldata.xlsx')
    all_wb.save(last_name)

    tkinter.messagebox.showinfo('提示', '合并完成,请查看《' + str(last_name) + '》')


def choose_xls_xlsx(max_row, all_sheet, filename, sheethead):
    """判断传进来的文件是xls还是xlsx，并使用对应的方法将数据写入汇总文件中"""
    if filename.endswith('.xls'):
        sub_wb = xlrd.open_workbook(filename)
        sub_sheet = sub_wb.sheet_by_index(0)

        for row in range(sub_sheet.nrows - sheethead):
            for column in range(sub_sheet.ncols):
                all_sheet.cell(max_row + row + 1, column + 1, sub_sheet.cell_value(row + sheethead, column))

    elif filename.endswith('.xlsx'):
        sub_wb = openpyxl.load_workbook(filename)
        sub_sheet = sub_wb.worksheets[0]

        for row in range(1, sub_sheet.max_row - sheethead + 1):
            for column in range(1, sub_sheet.max_column + 1):
                all_sheet.cell(max_row + row, column, sub_sheet.cell(row=row + sheethead, column=column).value)


def data_files():
    """提取当前目录下需要合并的文件名，放入列表中"""
    dadafiles = []
    for file in os.listdir('.'):
        if (file[-4:] == '.xls' or file[-5:] == '.xlsx') and file[:7] != 'alldata':
            dadafiles.append(file)
    return dadafiles


def treeview_sort_column(tv, col, reverse):
    """Treeview点击字段排序"""
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    l.sort(key=lambda t: (t[0]), reverse=reverse)

    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


"""DBF4"""

"""in ('')"""


def filename_in():
    """判断选择的文件是否为Excel"""
    fname_in.set('')
    address = filedialog.askopenfilename()
    if address:
        fname_in.set(address)

    if os.path.basename(address) and not (
            os.path.basename(address).endswith('.xls') or os.path.basename(address).endswith('.xlsx')):
        tkinter.messagebox.showinfo('提示', '所选文件格式不正确！')


def parts_of_choose(choose):
    """
    将输入的范围用正则拆开放入列表
    输入:  A1:A10,C1:C10,F1:F10
    return ["'A1':'A10'", "'C1':'C10'", "'F1':'F10'"]
    """
    result = {}
    pattern1 = r'[，|,]'  # 定义分隔符
    pattern2 = r'[：|:]'
    first_extraction = re.split(pattern1, choose)  # ['A1:A10', 'C1:C10', 'F1:F10']
    for i in first_extraction:
        second_extraction = re.split(pattern2, i)
        result[second_extraction[0]] = second_extraction[-1]

    return result


def get_data_from_excel(dir, choose, row=1):
    fname = os.path.basename(dir)
    if fname.endswith(".xls"):
        tkinter.messagebox.showinfo('提示', '请将文件转换成".xlsx"格式')

    elif fname.endswith(".xlsx"):
        wb = openpyxl.load_workbook(dir)
        sheet = wb.active
        result = []
        for v, k in choose.items():
            for row in sheet[v:k]:
                for cell in row:
                    if cell.value:
                        result.append(cell.value)
        # 调试
        # print("('", end="")
        # for i in result[:-1]:
        #     print(i, end="','")
        # print(str(result[-1]) + "')")
        wb.close()
        return result


def begin_format():
    format_in_resultText.delete(0.0, "end")
    dir = fname_in.get()
    choose = format_in_input.get()
    result = get_data_from_excel(dir, parts_of_choose(choose))
    format_in_resultText.insert("end", "('" + str(result[0]) + "','")
    for i in result[1:-1]:
        format_in_resultText.insert("end", str(i) + "','")
    format_in_resultText.insert("end", str(result[-1]) + "')")


"""获取群成员"""


def RElist_to_str(RElist):
    """用正则表达式获取的电话号码是以['18686622933']格式保存的，如果电话不存在则返回[]，要对list进行处理"""
    if RElist:
        return RElist[0]
    return ''


def user_sex(sex):
    """对微信'Sex'的值进行处理，0为未选择性别"""
    if sex == 1:
        return '男'
    elif sex == 2:
        return '女'
    else:
        return '-'


def login_wechat():
    """登陆微信，每次点击会先注销，再重新扫码登陆，并将微信群里列表放到Treeview中，包括活跃群和保存到通讯录中的群"""
    global grouplist
    itchat.logout()  # 退出已经登陆的微信，每次点击都重新扫码
    itchat.auto_login(enableCmdQR=False)
    # itchat.auto_login(enableCmdQR=False, hotReload=True)  # hotReload=True缓存登陆

    # 用于清空treeview中内容
    x = wchat_getlist_trv.get_children()
    for item in x:
        wchat_getlist_trv.delete(item)

    # 获取群信息，并写入treeview
    grouplist = []
    num = 1
    for i in itchat.get_chatrooms():
        ii = []
        ii.append(num)
        ii.append(i['NickName'])
        ii.append(i['MemberCount'])
        grouplist.append(ii)
        num += 1
    # print(grouplist)
    for s in grouplist:
        wchat_getlist_trv.insert('', s[0], text=s[0], values=(s[0], del_emoji(s[1]), s[2:]))

    # 更改按钮名称
    wchat_getlist_button1_name.set('登陆成功，点击重新登录')


def check_field():
    """选择需要输出的字段"""
    fiedls = {'个人微信名称': 0, '群昵称': 0, '群昵称电话': 0, '性别': 0}
    if wchat_getlist_checkb1 == 1:
        fiedls['个人微信名称'] = 1
    if wchat_getlist_checkb2 == 1:
        fiedls['群昵称'] = 1
    if wchat_getlist_checkb3 == 1:
        fiedls['群昵称电话'] = 1
    if wchat_getlist_checkb4 == 1:
        fiedls['性别'] = 1


def get_namelist(event):
    """双击Treeview选择群，将群成员名单输入到新弹出的Treeview中，这里因为emoji表情的原因不能取一层Treeview的群名称，要用索引选择"""
    global all_info, group_name, column_ch
    for i in wchat_getlist_trv.selection():
        group_index = wchat_getlist_trv.item(i, "values")[0]
        group_name = grouplist[int(group_index) - 1][1]
        search_result = itchat.search_chatrooms(name=group_name)[0]['UserName']
        user_info = itchat.update_chatroom(search_result, detailedMember=True)
        # print(search_result)
        row = 1
        all_info = []
        for ii in user_info['MemberList']:
            # print(ii, type(ii))
            one_user = []
            one_user.append(row)
            one_user.append(del_emoji(ii['NickName']))
            one_user.append(ii['DisplayName'])
            one_user.append(RElist_to_str(re.findall(r"1\d{10}", ii['DisplayName'])))
            one_user.append(user_sex(ii['Sex']))
            one_user.append(ii['Province'] + "-" + ii['City'])
            row += 1
            all_info.append(one_user)

        print(all_info)

        sub_root = tk.Tk()
        sub_root.title('"%s"群成员列表' % del_emoji(group_name))
        width, heigth = 700, 350
        sub_root.geometry('%dx%d+%d+%d' % (width, heigth, (sub_root.winfo_screenwidth() - width) / 2,
                                           (sub_root.winfo_screenheight() - heigth) / 2))

        sub_frame = tk.Frame(sub_root)
        sub_frame.pack(fill='both', expand=1)
        column_list = ('number', 'nickname', 'displayname', 'phonenum', 'sex', 'location')
        column_ch = ('序号', '微信名称', '群昵称', '手机号', '性别', '城市')
        sub_treeview = ttk.Treeview(sub_frame, height=15, show="headings", column=column_list)
        for col in column_list:
            sub_treeview.heading(col, text=col, command=lambda c=col: treeview_sort_column(sub_treeview, c, False))
        sub_treeview.pack(fill='both', expand=1, side=tk.LEFT, pady=2)
        for i in range(len(column_list)):
            sub_treeview.heading(column_list[i], text=column_ch[i])

        sub_treeview.column('number', width=40)  # 设置字段宽度
        sub_treeview.column('nickname', width=120)
        sub_treeview.column('displayname', width=200)
        sub_treeview.column('phonenum', width=120)
        sub_treeview.column('sex', width=40)
        sub_treeview.column('location', width=130)
        sub_vbar = ttk.Scrollbar(sub_frame, orient=tk.VERTICAL, command=sub_treeview.yview)  # 设置滚动条控件
        sub_vbar.pack(side=tk.LEFT, fill='y')  # 滚动条控件与被控制的Treeview同在一个容器中，并列放置，纵向填充
        sub_treeview.configure(yscrollcommand=sub_vbar.set)
        sub_button = tk.Button(sub_root, text='输出到Excel', font=myfont, command=to_excel, width=40).pack(pady=3)
        for s in all_info:
            sub_treeview.insert('', s[0], text=s[0], values=s[0:])
        sub_root.mainloop()


def to_excel():
    """将群成员名单输出到Excel中"""
    global all_info, group_name, column_ch
    wb = openpyxl.Workbook()
    worksheet = wb.active
    worksheet.title = '"%s"群成员名单' % group_name
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 5
    worksheet.column_dimensions['F'].width = 20

    cnum_field = 1
    for col in column_ch:
        worksheet[get_column_letter(cnum_field) + "1"] = col
        cnum_field += 1

    rnum = 2
    for row in all_info:
        for cnum in range(len(column_ch)):
            worksheet[get_column_letter(cnum + 1) + str(rnum)] = row[cnum]
        rnum += 1

    last_name = file_name('%s-成员名单.xlsx' % group_name)
    wb.save(last_name)
    if tkinter.messagebox.askyesno("提示", "结果已输出到《" + last_name + "》\n是否打开文件？"):
        try:
            os.system(last_name)
        except:
            pass
    else:
        pass


"""微信控制电脑摄像头拍照"""


def wchat_camera_button():
    sendMsg = u"{消息助手}：暂时无法回复"
    usageMsg = u"查看电脑前是谁请输入：cap，即可获取对方照片！"
    flag = 0
    nowTime = time.localtime()
    filename = str(nowTime.tm_mday) + str(nowTime.tm_hour) + str(nowTime.tm_min) + str(nowTime.tm_sec) + ".txt"
    myfile = open(filename, 'w')

    @itchat.msg_register('Text')
    def text_reply(msg):
        global flag
        message = msg['Text']
        fromName = msg['FromUserName']
        toName = msg['ToUserName']

        if toName == "filehelper":
            if message == "cap":
                cap = cv2.VideoCapture(0)
                ret, img = cap.read()
                cv2.imwrite("weixinTemp.jpg", img)
                itchat.send('@img@%s' % u'weixinTemp.jpg', 'filehelper')
                cap.release()

        elif flag == 1:
            itchat.send(sendMsg, fromName)
            myfile.write(message)
            myfile.write("\n")
            myfile.flush()

    if __name__ == '__main__':
        itchat.auto_login()
        itchat.send(usageMsg, "filehelper")
        itchat.run()


"""创建GUI"""
root = tk.Tk()
root.title('Work Support')
width, heigth = 500, 400
root.geometry(
    '%dx%d+%d+%d' % (width, heigth, (root.winfo_screenwidth() - width) / 2, (root.winfo_screenheight() - heigth) / 2))

"""创建顶级菜单"""
menubar = tk.Menu(root)
root.config(menu=menubar)
excelmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='Excel操作', menu=excelmenu)
datamenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='数据格式', menu=datamenu)
osmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='系统操作', menu=osmenu)
wxmenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='微信操作', menu=wxmenu)
othermenu = tk.Menu(menubar, tearoff=0)
menubar.add_cascade(label='其他', menu=othermenu)

"""创建下拉菜单以及对应功能的底层容器"""
excelmenu.add_command(label='数据对比', command=data_comparison)
data_comparison_frm0 = tk.Frame(root)
excelmenu.add_command(label='合并Excel', command=merge_excel)
merge_excel_frm0 = tk.Frame(root)
excelmenu.add_command(label='非标数据合并', command=non_standard)
non_standard_frm0 = tk.Frame(root)
datamenu.add_command(label='DBF4', command=format_DBF)
format_DBF_frm0 = tk.Frame(root)
datamenu.add_command(label="in('')", command=format_in)
format_in_frm0 = tk.Frame(root)
osmenu.add_command(label='定时关机', command=timing_shutdown)
osmenu.add_command(label='定时关闭IU', command=shutdown_iu)
wxmenu.add_command(label='天气预报', command=timing_shutdown)
wxmenu.add_command(label='获取群成员', command=wchat_getlist)
wchat_getlist_frm0 = tk.Frame(root)
wxmenu.add_command(label='控制摄像头', command=wchat_camera)
wchat_camera_frm0 = tk.Frame(root)
wxmenu.add_command(label='退出微信登录', command=logout_wchat)
othermenu.add_command(label='关于', command=about_this)
about_this_frm0 = tk.Frame(root)
othermenu.add_command(label='帮助', command=help_for)

switch = merge_excel_frm0

"""创建功能控件"""

"""数据对比"""
data_comparison_frm1 = tk.Frame(data_comparison_frm0, pady=1)
data_comparison_frm1.pack(fill=tk.X)
data_comparison_text0 = tk.Label(data_comparison_frm1, text='数据对比', font=myfont3).pack()
data_comparison_text = tk.Label(data_comparison_frm1,
                                text="\n提示:\n1、将要对比的两个表放在同一个excel中的前两位；\n2、两个表的字段顺序要相同。\n",
                                font=myfont, justify='left').pack()
fname = tk.StringVar()
data_comparison_ent = tk.Entry(data_comparison_frm1,
                               width=40, font=myfont, textvariable=fname).pack(fill=tk.X, expand=1, side=tk.LEFT)
data_comparison_button = tk.Button(data_comparison_frm1,
                                   width=10, text='选择文件', font=myfont, command=filename).pack(fill=tk.X, side=tk.LEFT)
data_comparison_frm2 = tk.Frame(data_comparison_frm0, pady=1)
data_comparison_frm2.pack(fill=tk.X)
data_comparison_input = tk.Entry(data_comparison_frm2, width=40, font=myfont)
data_comparison_input.pack(fill=tk.X, expand=1, side=tk.LEFT)
data_comparison_input.bind('<Double-1>', input_num)
data_comparison_input.bind('<1>', button3_disable)
data_comparison_input.bind('<Return>', begin_comparison)
data_comparison_button2 = tk.Button(data_comparison_frm2, width=10, text='开始对比', font=myfont,
                                    command=begin_comparison).pack(fill=tk.X, side=tk.LEFT)
data_comparison_frm3 = tk.Frame(data_comparison_frm0, pady=1)
data_comparison_frm3.pack(fill=tk.X)
data_comparison_canvas = tk.Canvas(data_comparison_frm3, width=400, height=22)  # , bg="white")
data_comparison_canvas.pack(fill=tk.X, expand=1, side=tk.LEFT)
data_comparison_button3 = tk.Button(data_comparison_frm3, width=10,
                                    text='输出结果', font=myfont, command=output_sheet)
data_comparison_button3.pack(fill=tk.X, side=tk.LEFT)
data_comparison_button3.configure(state='disable')

listhead = ('序号', '问题', '第几行', '内容')
data_comparison_lf = tk.Frame(data_comparison_frm0, pady=20)
data_comparison_lf.pack(fill='both', expand=1, padx=3)
data_comparison_tv = ttk.Treeview(data_comparison_lf, height=15, show="headings", columns=listhead)
for col in listhead:
    data_comparison_tv.heading(col, text=col, command=lambda c=col: treeview_sort_column(data_comparison_tv, c, False))

data_comparison_tv.pack(fill='both', expand=1, side=tk.LEFT)
data_comparison_tv.heading('序号', text='序号')
data_comparison_tv.heading('问题', text='问题')  # 设置字段的显示名称
data_comparison_tv.heading('第几行', text='第几行')
data_comparison_tv.heading('内容', text='内容')
data_comparison_tv.column('序号', width=40)
data_comparison_tv.column('问题', width=100)
data_comparison_tv.column('第几行', width=40)
data_comparison_tv.column('内容', width=300)
vbar = ttk.Scrollbar(data_comparison_lf, orient=tk.VERTICAL, command=data_comparison_tv.yview)  # 设置滚动条控件
vbar.pack(side=tk.LEFT, fill='y')  # 滚动条控件与被控制的Treeview同在一个容器中，并列放置，纵向填充
data_comparison_tv.configure(yscrollcommand=vbar.set)

"""合并Excel"""
merge_excel_frm1 = tk.Frame(merge_excel_frm0)
merge_excel_frm1.pack(fill='both', pady=10)
merge_excel_text0 = tk.Label(merge_excel_frm1, text='合并Excel', font=myfont3).pack()
merge_excel_text = tk.Label(merge_excel_frm1,
                            text="\n提示:\n1、将需要合并的文件放在同一文件夹下；\n2、数据表放在sheet1；\n3、文件表头行数要相等。\n",
                            font=myfont, justify='left').pack()
fdir = tk.StringVar()
merge_excel_ent = tk.Entry(merge_excel_frm1, width=40, font=myfont, textvariable=fdir).pack(fill=tk.X, expand=1,
                                                                                            side=tk.LEFT)
merge_excel_button = tk.Button(merge_excel_frm1,
                               width=10, text='选择文件夹', font=myfont, command=filedir).pack(fill=tk.X, side=tk.LEFT)
flist = tk.StringVar()
# flist.set((''))
merge_excel_listbox = tk.Listbox(merge_excel_frm0, listvariable=flist).pack(fill='both', expand=1)
merge_excel_frm2 = tk.Frame(merge_excel_frm0)
merge_excel_frm2.pack()
merge_excel_input = tk.Entry(merge_excel_frm2, width=40, font=myfont)
merge_excel_input.pack(fill=tk.X, side=tk.LEFT, pady=10)
merge_excel_input.bind('<Double-1>', input_num)
merge_excel_button2 = tk.Button(merge_excel_frm2, width=10, text='开始合并', font=myfont,
                                command=lambda: agg_excel(merge_excel_input, path)).pack(fill=tk.X, side=tk.LEFT)

"""DBF4"""

"""in ('')"""
format_in_frm1 = tk.Frame(format_in_frm0, pady=1)
format_in_frm1.pack(fill=tk.X)
format_in_text0 = tk.Label(format_in_frm1, text='in ('')', font=myfont3).pack()
format_in_text = tk.Label(format_in_frm1,
                          text="\n提示:\n1、将要对比的两个表放在同一个excel中的前两位；\n2、两个表的字段顺序要相同。\n",
                          font=myfont, justify='left').pack()
fname_in = tk.StringVar()
format_in_ent = tk.Entry(format_in_frm1,
                         width=40, font=myfont, textvariable=fname_in).pack(fill=tk.X, expand=1, side=tk.LEFT)
format_in_button = tk.Button(format_in_frm1,
                             width=10, text='选择文件', font=myfont, command=filename_in).pack(fill=tk.X, side=tk.LEFT)
format_in_frm2 = tk.Frame(format_in_frm0, pady=1)
format_in_frm2.pack(fill=tk.X)
format_in_input = tk.Entry(format_in_frm2, width=40, font=myfont)
format_in_input.pack(fill=tk.X, expand=1, side=tk.LEFT)
format_in_input.bind('<Double-1>', input_num)
format_in_input.bind('<1>', button3_disable)
format_in_button2 = tk.Button(format_in_frm2, width=10, text='格式化', font=myfont,
                              command=begin_format).pack(fill=tk.X, side=tk.LEFT)
format_in_frm3 = tk.Frame(format_in_frm0, padx=10, pady=20)
format_in_frm3.pack(fill='both', expand=1)
format_in_resultText = tk.Text(format_in_frm3, pady=1)
format_in_resultText.pack(fill='both', expand=1, side=tk.LEFT)

format_in_vbar = ttk.Scrollbar(format_in_frm3, orient=tk.VERTICAL, command=format_in_resultText.yview)  # 设置滚动条控件
format_in_vbar.pack(side=tk.LEFT, fill='y')
format_in_resultText.configure(yscrollcommand=format_in_vbar.set)

"""获取群成员"""
wchat_getlist_frm1 = tk.Frame(wchat_getlist_frm0)
wchat_getlist_frm1.pack()
wchat_getlist_text0 = tk.Label(wchat_getlist_frm1, text='获取微信群成员信息', font=myfont3).pack()
wchat_getlist_text = tk.Label(wchat_getlist_frm1,
                              text="\n提示:\n1、将聊天群保存到通讯录；\n2、扫码登陆微信；\n3、双击群名称。\n",
                              font=myfont, justify='left').pack()
wchat_getlist_button1_name = tk.StringVar()
wchat_getlist_button1_name.set("登陆微信")
wchat_getlist_button1 = tk.Button(wchat_getlist_frm1, textvariable=wchat_getlist_button1_name, font=myfont, width=40,
                                  command=login_wechat).pack(fill=tk.X, expand=1, pady=3)

wchat_getlist_frm12 = tk.Frame(wchat_getlist_frm0)
wchat_getlist_frm3 = tk.Frame(wchat_getlist_frm12)
wchat_getlist_frm3.pack(fill=tk.Y, expand=1, side=tk.LEFT)
wchat_getlist_lableframe = tk.LabelFrame(wchat_getlist_frm3, text='要获的信息', width=500)
wchat_getlist_lableframe.pack(fill='both', expand=1)
wchat_getlist_var1 = tk.IntVar()
wchat_getlist_var2 = tk.IntVar()
wchat_getlist_var3 = tk.IntVar()
wchat_getlist_var4 = tk.IntVar()
wchat_getlist_checkb1 = tk.Checkbutton(wchat_getlist_lableframe, text='个人微信名称', variable=wchat_getlist_var1,
                                       onvalue=1, offvalue=0, command=check_field)
wchat_getlist_checkb1.pack(anchor=tk.W)
wchat_getlist_checkb2 = tk.Checkbutton(wchat_getlist_lableframe, text='群昵称', variable=wchat_getlist_var2,
                                       onvalue=1, offvalue=0, command=check_field).pack(anchor=tk.W)
wchat_getlist_checkb3 = tk.Checkbutton(wchat_getlist_lableframe, text='群昵称电话', variable=wchat_getlist_var3,
                                       onvalue=1, offvalue=0, command=check_field).pack(anchor=tk.W)
wchat_getlist_checkb4 = tk.Checkbutton(wchat_getlist_lableframe, text='性别', variable=wchat_getlist_var4,
                                       onvalue=1, offvalue=0, command=check_field).pack(anchor=tk.W)
wchat_getlist_frm3.forget()
wchat_getlist_frm2 = tk.LabelFrame(wchat_getlist_frm12, text='群列表(双击查看成员名单)：')
wchat_getlist_frm12.pack(fill='both', expand=1, padx=2, pady=10)
wchat_getlist_frm2.pack(fill='both', expand=1, side=tk.LEFT, padx=1)
wchat_getlist_trv = ttk.Treeview(wchat_getlist_frm2, height=10, show="headings",
                                 columns=('Number', 'NickName', 'MemberCount'))
wchat_getlist_trv.bind('<Double-Button-1>', get_namelist)
# treeview字段点击排序
for col in ('Number', 'NickName', 'MemberCount'):
    wchat_getlist_trv.heading(col, text=col, command=lambda c=col: treeview_sort_column(wchat_getlist_trv, c, False))
wchat_getlist_trv.pack(fill='both', expand=1, side=tk.LEFT)
wchat_getlist_trv.heading('Number', text='序号')  # 设置字段的显示名称
wchat_getlist_trv.heading('NickName', text='群聊名称')
wchat_getlist_trv.heading('MemberCount', text='人数')
wchat_getlist_trv.column('Number', width=30)
wchat_getlist_trv.column('NickName', width=250)
wchat_getlist_trv.column('MemberCount', width=40)
wchat_getlist_vbar = ttk.Scrollbar(wchat_getlist_frm2, orient=tk.VERTICAL, command=wchat_getlist_trv.yview)  # 设置滚动条控件
wchat_getlist_vbar.pack(side=tk.LEFT, fill='y')
wchat_getlist_trv.configure(yscrollcommand=wchat_getlist_vbar.set)
# wchat_getlist_button2 = tk.Button(wchat_getlist_frm0, text='获取群成员名单', font=myfont, width=40,
#                                   command=get_namelist).pack(pady=15)


"""微信控制电脑摄像头拍照"""
wchat_camera_text0 = tk.Label(wchat_camera_frm0, text='微信控制电脑摄像头', font=myfont3).pack()
wchat_camera_text = tk.Label(wchat_camera_frm0, text='\n提示:\n1、扫码登陆微信；\n2、在文件传输助手中发送"cap"获取图像。\n',
                             font=myfont, justify='left').pack()
wchat_camera_button = tk.Button(wchat_camera_frm0, text="登陆微信", font=myfont, width=40, command=wchat_camera_button)
wchat_camera_button.pack(fill=tk.X, expand=1, pady=3)

"""关于"""
about_this_frm1 = tk.Frame(about_this_frm0)
about_this_frm1.pack(fill='both', expand=1)
about_this_lb = tk.Label(about_this_frm1, text="版本号：0.15\n\n作者：陈博文", font=myfont).pack(fill='both', expand=1)

root.mainloop()
